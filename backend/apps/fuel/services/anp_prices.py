"""
Service to fetch fuel prices from ANP (Agência Nacional do Petróleo).

The ANP publishes weekly price surveys in XLSX format at:
https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/levantamento-de-precos-de-combustiveis-ultimas-semanas-pesquisadas
"""
import io
import logging
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Optional

import requests
from django.utils import timezone
from openpyxl import load_workbook

from apps.core.models import FuelType
from apps.fuel.models import FuelPriceSnapshot, FuelPriceSource

logger = logging.getLogger(__name__)

ANP_BASE_URL = "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/arquivos-lpc"

# Mapping of ANP product names to our FuelType
ANP_FUEL_MAPPING = {
    'GASOLINA COMUM': FuelType.GASOLINE,
    'GASOLINA C COMUM': FuelType.GASOLINE,
    'ETANOL HIDRATADO': FuelType.ETHANOL,
    'ETANOL HIDRATADO COMUM': FuelType.ETHANOL,
    'ÓLEO DIESEL': FuelType.DIESEL,
    'ÓLEO DIESEL S10': FuelType.DIESEL,
    'DIESEL S10': FuelType.DIESEL,
}


def get_anp_file_url(date: Optional[datetime] = None) -> str:
    """
    Generate the URL for the ANP weekly summary file.

    Files follow pattern: resumo_semanal_lpc_YYYY-MM-DD_YYYY-MM-DD.xlsx
    where dates are start/end of the survey week (Sunday to Saturday).
    """
    if date is None:
        date = timezone.now()

    # Find the most recent Saturday (end of survey week)
    days_since_saturday = (date.weekday() + 2) % 7
    end_date = date - timedelta(days=days_since_saturday)
    start_date = end_date - timedelta(days=6)

    year = end_date.year
    start_str = start_date.strftime('%Y-%m-%d')
    end_str = end_date.strftime('%Y-%m-%d')

    return f"{ANP_BASE_URL}/{year}/resumo_semanal_lpc_{start_str}_{end_str}.xlsx"


def download_anp_file(url: str) -> Optional[bytes]:
    """Download the ANP XLSX file."""
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        return response.content
    except requests.RequestException as e:
        logger.error(f"Failed to download ANP file from {url}: {e}")
        return None


def parse_anp_xlsx(content: bytes) -> dict[str, Decimal]:
    """
    Parse the ANP XLSX file and extract national average prices.

    Returns a dict mapping FuelType to average price per liter.
    """
    prices = {}

    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)

        # The file typically has a sheet with national averages
        # Look for sheets with "BRASIL" or "RESUMO" in the name
        target_sheets = []
        for sheet_name in wb.sheetnames:
            sheet_upper = sheet_name.upper()
            if 'BRASIL' in sheet_upper or 'RESUMO' in sheet_upper or 'MÉDIA' in sheet_upper:
                target_sheets.append(sheet_name)

        if not target_sheets:
            # Try the first sheet
            target_sheets = [wb.sheetnames[0]]

        for sheet_name in target_sheets:
            ws = wb[sheet_name]

            # Find column indices for product name and average price
            header_row = None
            product_col = None
            price_col = None

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                if row is None:
                    continue

                for col_idx, cell in enumerate(row):
                    if cell is None:
                        continue
                    cell_str = str(cell).upper().strip()

                    if 'PRODUTO' in cell_str or 'COMBUSTÍVEL' in cell_str:
                        header_row = row_idx
                        product_col = col_idx
                    elif 'MÉDIA' in cell_str and 'REVENDA' in cell_str:
                        price_col = col_idx
                    elif 'PREÇO MÉDIO' in cell_str:
                        price_col = col_idx

                if product_col is not None and price_col is not None:
                    break

            if product_col is None or price_col is None:
                logger.warning(f"Could not find required columns in sheet {sheet_name}")
                continue

            # Read data rows
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if row is None or len(row) <= max(product_col, price_col):
                    continue

                product = row[product_col]
                price = row[price_col]

                if product is None or price is None:
                    continue

                product_str = str(product).upper().strip()

                # Match product to our fuel types
                for anp_name, fuel_type in ANP_FUEL_MAPPING.items():
                    if anp_name in product_str:
                        try:
                            # Handle both float and string prices
                            if isinstance(price, (int, float)):
                                price_decimal = Decimal(str(price)).quantize(Decimal('0.0001'))
                            else:
                                # Handle Brazilian format (comma as decimal separator)
                                price_str = str(price).replace('.', '').replace(',', '.')
                                price_decimal = Decimal(price_str).quantize(Decimal('0.0001'))

                            # Only keep if we don't have this fuel type yet or if this is higher quality
                            if fuel_type not in prices:
                                prices[fuel_type] = price_decimal
                                logger.info(f"Found price for {fuel_type}: R$ {price_decimal}")
                        except (InvalidOperation, ValueError) as e:
                            logger.warning(f"Could not parse price '{price}' for {product}: {e}")
                        break

        wb.close()

    except Exception as e:
        logger.error(f"Error parsing ANP XLSX: {e}")

    return prices


def fetch_and_save_anp_prices() -> dict[str, any]:
    """
    Main function to fetch ANP prices and save to database.

    Returns a dict with status and details.
    """
    result = {
        'success': False,
        'prices_updated': [],
        'errors': [],
    }

    # Try current week first, then previous week
    urls_to_try = [
        get_anp_file_url(),
        get_anp_file_url(timezone.now() - timedelta(days=7)),
    ]

    content = None
    used_url = None

    for url in urls_to_try:
        logger.info(f"Trying to fetch ANP data from: {url}")
        content = download_anp_file(url)
        if content:
            used_url = url
            break

    if not content:
        result['errors'].append("Could not download ANP file from any URL")
        return result

    # Parse the XLSX
    prices = parse_anp_xlsx(content)

    if not prices:
        result['errors'].append("No prices found in ANP file")
        return result

    # Save to database
    now = timezone.now()

    for fuel_type, price in prices.items():
        try:
            snapshot, created = FuelPriceSnapshot.objects.update_or_create(
                fuel_type=fuel_type,
                station=None,  # National average
                source=FuelPriceSource.EXTERNAL_ANP,
                defaults={
                    'price_per_liter': price,
                    'collected_at': now,
                }
            )

            action = 'created' if created else 'updated'
            result['prices_updated'].append({
                'fuel_type': fuel_type,
                'price': str(price),
                'action': action,
            })
            logger.info(f"ANP price {action}: {fuel_type} = R$ {price}")

        except Exception as e:
            error_msg = f"Failed to save price for {fuel_type}: {e}"
            result['errors'].append(error_msg)
            logger.error(error_msg)

    result['success'] = len(result['prices_updated']) > 0
    result['source_url'] = used_url

    return result
